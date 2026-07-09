from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math

# ── 資料 ──────────────────────────────────────────────
PRODUCTS = [
  # 美麗多
  {"id":"MLMN005","name":"美麗多木質墊料(絲)0.05公斤","cat":"美麗多","stock":516},
  {"id":"MLMP07","name":"美麗多木質墊料(大顆粒)-3.5L","cat":"美麗多","stock":-60},
  {"id":"MLMP10","name":"美麗多木質墊料(大顆粒)-50L","cat":"美麗多","stock":-50},
  {"id":"MLMP2","name":"美麗多木質墊料(大顆粒)-10L","cat":"美麗多","stock":2627},
  {"id":"MLMP4","name":"美麗多木質墊料(大顆粒)-20L","cat":"美麗多","stock":-22},
  {"id":"MLMS07","name":"美麗多木質墊料(小顆粒)-3.5L","cat":"美麗多","stock":25},
  {"id":"MLMS10","name":"美麗多木質墊料(小顆粒)-50L","cat":"美麗多","stock":25},
  {"id":"MLMS2","name":"美麗多木質墊料(小顆粒)-10L","cat":"美麗多","stock":3004},
  {"id":"MMC5","name":"美麗多木刨花(小)-25L","cat":"美麗多","stock":543},
  {"id":"MMCP5","name":"美麗多木刨花(大)-25L","cat":"美麗多","stock":1342},
  {"id":"MMN50G","name":"美麗多純淨白楊木絲築巢(8入)","cat":"美麗多","stock":0},
  {"id":"MMR10L","name":"美麗多實驗室無塵純萃天然木梗","cat":"美麗多","stock":510},
  # 飼料
  {"id":"ALC1","name":"摩米苜蓿草磚","cat":"飼料","stock":867},
  {"id":"ALC5","name":"摩米苜蓿草磚50磅","cat":"飼料","stock":0},
  {"id":"APPLE","name":"摩米特級凍乾蘋果小食 15克裝","cat":"飼料","stock":1030},
  {"id":"BANANA","name":"摩米特級凍乾香蕉小食 15克裝","cat":"飼料","stock":301},
  {"id":"BERRY","name":"摩米特級凍乾草莓小食 15克裝","cat":"飼料","stock":1220},
  {"id":"BLUEBERRY","name":"摩米特級凍乾藍莓小食 15克裝","cat":"飼料","stock":999},
  {"id":"MCCA64","name":"摩米營養護極幼草粉 - 原味 64克裝","cat":"飼料","stock":438},
  {"id":"MCCATB","name":"摩米草粉試吃禮-原味","cat":"飼料","stock":-1132},
  {"id":"MCCATB-S","name":"摩米草粉自用(單條)-原味","cat":"飼料","stock":485},
  {"id":"MCCB64","name":"摩米營養護極幼草粉 - 香蕉味 64克裝","cat":"飼料","stock":-10},
  {"id":"MCCBB100","name":"摩米保康全營養草粉-香蕉味(大容量)*100條","cat":"飼料","stock":-3},
  {"id":"MCCBTB","name":"摩米草粉試吃禮-香蕉味","cat":"飼料","stock":-546},
  {"id":"MCCBTB-S","name":"摩米草粉自用(單條)-香蕉味","cat":"飼料","stock":964},
  {"id":"MKHF6","name":"摩米卡莉寵物飲水機專用濾芯","cat":"飼料","stock":508},
  {"id":"MKHSTW","name":"寵物飲水機-標準版-蘇菲白","cat":"飼料","stock":156},
  {"id":"PAPAYA","name":"摩米特級凍乾木瓜小食 15克裝","cat":"飼料","stock":0},
  {"id":"PEACH","name":"摩米特級凍乾水蜜桃小食 15克裝","cat":"飼料","stock":566},
  {"id":"PEAR","name":"摩米特級凍乾雪梨小食 15克裝","cat":"飼料","stock":884},
  {"id":"PINEAPPLE","name":"摩米特級凍乾鳯梨小食 15克裝","cat":"飼料","stock":-56},
  {"id":"TIMC1","name":"摩米提摩西草磚","cat":"飼料","stock":414},
  {"id":"TIMC50","name":"摩米提摩西草磚50磅","cat":"飼料","stock":0},
  {"id":"TIMACTB","name":"苜蓿草磚試吃包*5","cat":"飼料","stock":-148},
  {"id":"TIMCA1","name":"摩米營養全幼兔 A 1公斤裝","cat":"飼料","stock":218},
  {"id":"TIMCA5","name":"摩米營養全幼兔 A 5公斤裝","cat":"飼料","stock":37},
  {"id":"TIMCG1","name":"摩米營養全天竺鼠 G 1公斤裝","cat":"飼料","stock":887},
  {"id":"TIMCG5","name":"摩米營養全天竺鼠 G 5公斤裝","cat":"飼料","stock":39},
  {"id":"TIMCT1","name":"摩米營養全成兔 T 1公斤裝","cat":"飼料","stock":1223},
  {"id":"TIMCT5","name":"摩米營養全成兔 T 5公斤裝","cat":"飼料","stock":214},
  {"id":"TIMIC1","name":"摩米營養全老兔 IC1公斤裝","cat":"飼料","stock":168},
  {"id":"TIMIC5","name":"摩米營養全老兔 IC 5公斤裝","cat":"飼料","stock":47},
  {"id":"TIMNT4","name":"摩米全天然T 4磅裝(1.8公斤)","cat":"飼料","stock":6},
  {"id":"TIMTCTB","name":"提摩西草磚試吃包*5","cat":"飼料","stock":-94},
  {"id":"TPCA","name":"飼料試吃-幼兔*10","cat":"飼料","stock":-3},
  {"id":"TPCG","name":"飼料試吃-天竺鼠*10","cat":"飼料","stock":-1},
  {"id":"TPCT","name":"飼料試吃-成兔*10","cat":"飼料","stock":-1},
  {"id":"TPIC","name":"飼料試吃-老兔*10","cat":"飼料","stock":-4},
  {"id":"TPNT","name":"飼料試吃-全天然*10","cat":"飼料","stock":2},
  # 牧草
  {"id":"ALFA1","name":"苜蓿草1公斤(未壓縮)","cat":"牧草","stock":216},
  {"id":"ALFA2","name":"苜蓿草2公斤(壓縮)","cat":"牧草","stock":-92},
  {"id":"ALFAZX","name":"苜蓿草(無壓縮)","cat":"牧草","stock":-130},
  {"id":"ALFTB-10","name":"苜蓿試吃包*10","cat":"牧草","stock":-2},
  {"id":"MOMI-ALFAZX","name":"苜蓿草(壓縮)","cat":"牧草","stock":-33},
  {"id":"MOMI-FIRST","name":"First Cut第一割提牧草(提摩西草)","cat":"牧草","stock":-131},
  {"id":"MOMI-SECOND","name":"Second Cut 第二割提牧草(提摩西草)","cat":"牧草","stock":-2217},
  {"id":"ORCGZX","name":"果園草(捆)","cat":"牧草","stock":-522},
  {"id":"PFTF2","name":"自用牧草-未壓一割","cat":"牧草","stock":12},
  {"id":"PFTO2","name":"自用牧草-果園草","cat":"牧草","stock":1},
  {"id":"PFTS2","name":"自用牧草-未壓二割","cat":"牧草","stock":-6},
  {"id":"PORCG","name":"農夫特級果園草1公斤","cat":"牧草","stock":15},
  {"id":"PTA2","name":"自用牧草-苜蓿","cat":"牧草","stock":-26},
  {"id":"PTF2","name":"自用牧草-一割","cat":"牧草","stock":16},
  {"id":"PTIMF1","name":"農夫特選第一割提摩西草1公斤","cat":"牧草","stock":12},
  {"id":"PTIMFTB-10","name":"皇牌一割試吃包*10","cat":"牧草","stock":56},
  {"id":"PTIMS1","name":"農夫特選第二割提摩西草1公斤","cat":"牧草","stock":78},
  {"id":"PTIMSTB-10","name":"皇牌二割試吃包*10","cat":"牧草","stock":72},
  {"id":"PTS2","name":"自用牧草-二割","cat":"牧草","stock":12},
  {"id":"TIMF","name":"First Cut第一割提牧草(未壓)","cat":"牧草","stock":-15},
  {"id":"TIMF05","name":"First第一割牧草(提摩西草) 500G","cat":"牧草","stock":25},
  {"id":"TIMF1","name":"First第一割牧草(提摩西草) 1000G","cat":"牧草","stock":277},
  {"id":"TIMF10","name":"First第一割提牧草(提摩西草) 10KG","cat":"牧草","stock":-37},
  {"id":"TIMF25","name":"First第一割牧草(提摩西草) 2.5KG","cat":"牧草","stock":236},
  {"id":"TIMF5","name":"First第一割提牧草(提摩西草) 5KG","cat":"牧草","stock":-12},
  {"id":"TIMFTB-10","name":"一割試吃包*10","cat":"牧草","stock":-30},
  {"id":"TIMS","name":"Second Cut 第二割提牧草(未壓)","cat":"牧草","stock":-503},
  {"id":"TIMS05","name":"Second第二割牧草(提摩西草) 500G","cat":"牧草","stock":60},
  {"id":"TIMS1","name":"Second第二割提牧草(提摩西草) 1000G","cat":"牧草","stock":204},
  {"id":"TIMS10","name":"Second第二割提牧草(提摩西草) 10KG","cat":"牧草","stock":-12},
  {"id":"TIMS25","name":"Second第二割提牧草(提摩西草) 2.5KG","cat":"牧草","stock":206},
  {"id":"TIMS5","name":"Second第二割提牧草(提摩西草) 5KG","cat":"牧草","stock":-7},
  {"id":"TIMSTB-10","name":"二割試吃包*10","cat":"牧草","stock":-73},
  {"id":"TIMTCTB2","name":"提摩西草磚試吃包*5","cat":"牧草","stock":-94},
]

SALES = {
  "MLMN005":[52,64,72],"MLMP07":[0,0,19],"MLMP10":[13,25,39],"MLMP2":[1323,1825,2295],
  "MLMP4":[69,298,305],"MLMS07":[84,189,122],"MLMS10":[19,33,19],"MLMS2":[1631,1447,1430],
  "MMC5":[58,62,102],"MMCP5":[90,101,159],"MMN50G":[0,0,6],"MMR10L":[19,32,99],
  "APPLE":[653,618,662],"BANANA":[224,190,219],"BERRY":[329,409,343],"BLUEBERRY":[265,323,197],
  "MCCA64":[275,292,317],"MCCATB":[136,100,77],"MCCATB-S":[71,40,181],"MCCB64":[294,296,370],
  "MCCBB100":[0,1,0],"MCCBTB":[136,70,64],"MCCBTB-S":[20,34,161],"MKHF6":[17,5,0],
  "MKHSTW":[20,12,0],"PAPAYA":[8,39,5],"PEACH":[130,183,138],"PEAR":[89,175,89],
  "PINEAPPLE":[160,134,263],"TIMCA1":[411,332,388],"TIMCA5":[7,17,8],"TIMCG1":[384,293,279],
  "TIMCG5":[19,26,6],"TIMCT1":[860,1243,726],"TIMCT5":[88,44,92],"TIMIC1":[332,212,289],
  "TIMIC5":[21,21,18],"TIMNT4":[43,21,17],"TPCA":[18,12,23],"TPCG":[18,16,19],
  "TPCT":[21,17,23],"TPIC":[18,12,23],"TPNT":[6,11,23],
  "ALC1":[366,431,506],"ALC5":[0,0,0],"TIMC1":[736,933,1139],"TIMC50":[0,0,0],
  "TIMACTB":[71,15,15],"TIMTCTB":[71,21,25],
  "ALFA1":[482,164,669],"ALFA2":[16,76,0],"ALFAZX":[30,0,20],"ALFTB-10":[0,2,0],
  "MOMI-ALFAZX":[0,0,3],"MOMI-FIRST":[0,0,10],"MOMI-SECOND":[0,0,675],"ORCGZX":[0,131,95],
  "PFTF2":[0,0,1],"PFTO2":[0,1,1],"PFTS2":[0,0,1],"PORCG":[61,50,123],"PTA2":[176,153,202],
  "PTF2":[1,40,87],"PTIMF1":[275,300,288],"PTIMFTB-10":[0,2,0],"PTIMS1":[288,222,458],
  "PTIMSTB-10":[0,2,0],"PTS2":[97,26,8],"TIMACTB":[71,15,15],"TIMF":[0,0,15],
  "TIMF05":[115,126,220],"TIMF1":[428,447,547],"TIMF10":[63,77,119],"TIMF25":[484,377,483],
  "TIMF5":[30,53,53],"TIMFTB-10":[27,17,17],"TIMS":[0,10,248],"TIMS05":[311,335,449],
  "TIMS1":[821,838,1005],"TIMS10":[162,85,149],"TIMS25":[707,1123,982],"TIMS5":[78,68,71],
  "TIMSTB-10":[72,17,10],"TIMTCTB2":[71,21,25],
}

LEAD = 60
TARGET = 15
SAFETY = 7

def rec_qty(p):
    s = SALES.get(p["id"], [0,0,0])
    avg = sum(s)/3 if s else 0
    daily = avg/30
    stock = p["stock"]
    if daily <= 0:
        return 0, 999, avg, daily
    days_left = stock / daily
    # need (LEAD+TARGET+SAFETY) days of supply from today
    order = math.ceil(max(0, (LEAD + TARGET + SAFETY) * daily - stock))
    return order, round(days_left,1), round(avg,1), round(daily,2)

def status(days):
    if days < 45: return "⚠️ 緊急"
    if days < 90: return "⚠ 留意"
    return "✅ 充足"

# ── 建立 Excel ─────────────────────────────────────────
wb = Workbook()

CAT_COLORS = {"美麗多":"8B5CF6","飼料":"F59E0B","牧草":"22C55E"}
CAT_TEXT   = {"美麗多":"FFFFFF","飼料":"FFFFFF","牧草":"FFFFFF"}
HEADER_BG  = "1E293B"
HEADER_FG  = "FFFFFF"
ROW_ALT    = "F8FAFC"
DANGER_BG  = "FEE2E2"
WARN_BG    = "FEF3C7"

thin = Side(style="thin", color="E2E8F0")
border = Border(left=thin,right=thin,top=thin,bottom=thin)

cats = ["美麗多","飼料","牧草"]
for ci, cat in enumerate(cats):
    ws = wb.active if ci==0 else wb.create_sheet()
    ws.title = cat
    prods = [p for p in PRODUCTS if p["cat"]==cat]

    # header
    headers = ["品號","品名","即時庫存","月均銷量","日均銷量","庫存天數","狀態","建議訂購量"]
    col_w   = [14,   36,    10,       10,       10,       10,      10,    12]
    for c,h in enumerate(headers,1):
        cell = ws.cell(row=1,column=c,value=h)
        cell.font      = Font(name="Arial",bold=True,color=HEADER_FG,size=10)
        cell.fill      = PatternFill("solid",start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(c)].width = col_w[c-1]
    ws.row_dimensions[1].height = 28

    for ri, p in enumerate(prods, 2):
        order, days_left, avg_m, daily = rec_qty(p)
        st = status(days_left) if daily>0 else "—"
        row = [p["id"], p["name"], p["stock"], round(avg_m,1), round(daily,2),
               days_left if daily>0 else "—", st, order if order>0 else "—"]
        bg = DANGER_BG if days_left<45 and daily>0 else (WARN_BG if days_left<90 and daily>0 else (ROW_ALT if ri%2==0 else "FFFFFF"))
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=c, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center" if c!=2 else "left", vertical="center")
            cell.border    = border
            if c==8 and isinstance(val,int) and val>0:
                cell.font = Font(name="Arial",size=10,bold=True,color="B45309")

    ws.freeze_panes = "A2"

# remove default sheet if extra
if len(wb.sheetnames)>3:
    del wb[wb.sheetnames[0]]

out = "/sessions/funny-upbeat-euler/mnt/outputs/MOMI訂購建議_20260709.xlsx"
wb.save(out)
print("saved:", out)
